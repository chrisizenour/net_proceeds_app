import pandas as pd
from io import BytesIO
from tempfile import NamedTemporaryFile
from datetime import datetime
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.dimensions import ColumnDimension, SheetDimension, SheetFormatProperties
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, numbers, DEFAULT_FONT
from openpyxl import drawing

def inputs_to_excel(
        agent,
        date,
        seller,
        address,
        first_trust,
        second_trust,
        annual_taxes,
        prorated_taxes,
        annual_hoa_condo_amt,
        prorated_annual_hoa_condo_amt,
        list_price,
        down_payment_pct,
        closing_subsidy_amt,
        listing_company_pct,
        selling_company_pct,
        processing_fee,
        settlement_fee,
        deed_preparation_fee,
        lien_release_fee,
        lien_trust_qty,
        recording_release_fee,
        recording_release_qty,
        grantors_tax_pct,
        congestion_tax_pct,
        pest_inspection_fee,
        poa_condo_disclosure_fee,
        other_fee_name,
        other_fee_amt,
):


    wb = Workbook()
    dest_filename = f"net_proceeds_form_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    ws1 = wb.active
    ws1.title = 'net_proceeds'
    ws1.print_area = 'B2:F56'
    ws1.set_printer_settings(paper_size=1, orientation='portrait')
    ws1.sheet_properties.pageSetUpPr.fitToPage = True

    white_fill = '00FFFFFF'
    yellow_fill = '00FFFF00'
    black_fill = '00000000'
    font_size = 12
    thick = Side(border_style='thick')
    thin = Side(border_style='thin')
    hair = Side(border_style='hair')
    DEFAULT_FONT.size = font_size
    ws1.column_dimensions['B'].width = 1.5
    ws1.column_dimensions['C'].width = 42.83
    ws1.column_dimensions['D'].width = 42.83
    ws1.column_dimensions['E'].width = 20.83
    ws1.column_dimensions['F'].width = 1.5

    acct_fmt = '_($* #,##0_);[Red]_($* (#,##0);_($* "-"??_);_(@_)'
    pct_fmt = '0.00%'

    for row in ws1.iter_rows(min_row=1, max_row=70, min_col=1, max_col=10):
        for cell in row:
            cell.fill = PatternFill(start_color=white_fill, end_color=white_fill, fill_type='solid')

    # Build Black Border
    ws1.merge_cells('A1:G1')
    top_left_border_one = ws1['A1']
    top_left_border_one.fill = PatternFill('solid', fgColor=black_fill)

    ws1.merge_cells('A2:A57')
    top_left_border_two = ws1['A2']
    top_left_border_two.fill = PatternFill('solid', fgColor=black_fill)

    ws1.merge_cells('B57:G57')
    top_left_border_three = ws1['B57']
    top_left_border_three.fill = PatternFill('solid', fgColor=black_fill)

    ws1.merge_cells('G2:G56')
    top_left_border_four = ws1['G2']
    top_left_border_four.fill = PatternFill('solid', fgColor=black_fill)

    # Build Header
    ws1.merge_cells('C3:E3')
    top_left_cell_one = ws1['C3']
    top_left_cell_one.value = 'Estimated Net Proceeds Based on Recommended List Price'
    top_left_cell_one.font = Font(bold=True)
    # top_left_cell_one.fill = PatternFill('solid', fgColor=white_fill)
    top_left_cell_one.alignment = Alignment(horizontal='center')

    ws1.merge_cells('C4:E4')
    top_left_cell_two = ws1['C4']
    top_left_cell_two.value = f'{seller} - {address}'
    top_left_cell_two.font = Font(bold=True)
    # top_left_cell_two.fill = PatternFill('solid', fgColor=white_fill)
    top_left_cell_two.alignment = Alignment(horizontal='center')

    ws1.merge_cells('C5:E5')
    top_left_cell_three = ws1['C5']
    top_left_cell_three.value = f'Date Prepared: {date}'
    top_left_cell_three.font = Font(bold=True)
    # top_left_cell_three.fill = PatternFill('solid', fgColor=white_fill)
    top_left_cell_three.alignment = Alignment(horizontal='center')

    d8 = ws1['D8']
    d8.value = 'Rec. List Price ($)'
    d8.font = Font(bold=True)
    d8.number_format = acct_fmt
    d8.alignment = Alignment(horizontal='right')

    e8 = ws1['E8']
    e8.value = list_price
    e8.font = Font(bold=True)
    e8.number_format = acct_fmt
    e8.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # d9 = ws1['D9']
    # d9.value = 'Down Pmt (%)'
    # d9.font = Font(bold=True)
    # d9.alignment = Alignment(horizontal='right')
    #
    # e9 = ws1['E9']
    # e9.value = down_payment_pct
    # e9.font = Font(bold=True)
    # e9.number_format = pct_fmt
    # e9.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Housing Related Cost Table
    c10 = ws1['C10']
    c10.value = 'HOUSING-RELATED COSTS'
    c10.font = Font(bold=True)
    c10.border = Border(bottom=thin)

    d10 = ws1['D10']
    d10.value = 'Calculation Description'
    d10.font = Font(bold=True)
    d10.border = Border(bottom=thin)

    c11 = ws1['C11']
    c11.value = 'Estimated Payoff - 1st Trust'
    c11.border = Border(top=thin, bottom=hair, left=thin)

    d11 = ws1['D11']
    d11.value = 'Principal Balance of Loan'
    d11.border = Border(top=thin, bottom=hair)

    e11 = ws1['E11']
    e11.value = first_trust
    e11.number_format = acct_fmt
    e11.border = Border(top=thin, right=thin, bottom=hair)

    c12 = ws1['C12']
    c12.value = 'Estimated Payoff - 2nd Trust'
    c12.border = Border(top=hair, bottom=hair, left=thin)

    d12 = ws1['D12']
    d12.value = 'Principal Balance of Loan'
    d12.border = Border(top=hair, bottom=hair)

    e12 = ws1['E12']
    e12.value = second_trust
    e12.number_format = acct_fmt
    e12.border = Border(top=hair, bottom=hair, right=thin)

    c13 = ws1['c13']
    c13.value = 'Purchaser Closing Cost / Contract'
    c13.border = Border(top=hair, bottom=hair, left=thin)

    d13 = ws1['D13']
    d13.value = 'Negotiated Into Contract'
    d13.border = Border(top=hair, bottom=hair)

    e13 = ws1['E13']
    e13.value = closing_subsidy_amt
    e13.number_format = acct_fmt
    e13.border = Border(top=hair, bottom=hair, right=thin)

    c14 = ws1['C14']
    c14.value = 'Prorated Taxes / Assessments'
    c14.border = Border(top=hair, bottom=hair, left=thin)

    d14 = ws1['D14']
    d14.value = '1 Year of Taxes Divided by 12 Multiplied by 3'
    d14.border = Border(top=hair, bottom=hair)

    e14 = ws1['E14']
    e14.value = prorated_taxes
    e14.number_format = acct_fmt
    e14.border = Border(top=hair, bottom=hair, right=thin)

    c15 = ws1['C15']
    c15.value = 'Prorated HOA / Condo Dues'
    c15.border = Border(top=hair, bottom=thin, left=thin)

    d15 = ws1['D15']
    d15.value = '1 Year of Dues Divided by 12 Multiplied by 3'
    d15.border = Border(top=hair, bottom=thin)

    e15 = ws1['E15']
    e15.value = prorated_annual_hoa_condo_amt
    e15.number_format = acct_fmt
    e15.border = Border(top=hair, bottom=thin, right=thin)

    e16 = ws1['E16']
    e16.value = '=SUM(E11:E15)'
    e16.font = Font(bold=True)
    e16.number_format = acct_fmt
    e16.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Broker and Financing Costs Table
    c18 = ws1['C18']
    c18.value = 'BROKERAGE & FINANCING COSTS'
    c18.font = Font(bold=True)
    c18.border = Border(bottom=thin)

    c19 = ws1['C19']
    c19.value = 'Listing Company Compensation'
    c19.border = Border(top=thin, bottom=hair, left=thin)

    d19 = ws1['D19']
    d19.value = '% from Listing Agreement * Rec. List Price'
    d19.border = Border(top=thin, bottom=hair)

    e19 = ws1['E19']
    e19.value = listing_company_pct * list_price
    e19.number_format = acct_fmt
    e19.border = Border(top=thin, bottom=hair, right=thin)

    c20 = ws1['C20']
    c20.value = 'Selling Company Compensation'
    c20.border = Border(top=hair, bottom=hair, left=thin)

    d20 = ws1['D20']
    d20.value = '% from Listing Agreement * Rec. List Price'
    d20.border = Border(top=hair, bottom=hair)

    e20 = ws1['E20']
    e20.value = selling_company_pct * list_price
    e20.number_format = acct_fmt
    e20.border = Border(top=hair, bottom=hair, right=thin)

    c21 = ws1['C21']
    c21.value = 'Processing Fee'
    c21.border = Border(top=hair, bottom=thin, left=thin)

    d21 = ws1['D21']
    d21.border = Border(top=hair, bottom=thin)

    e21 = ws1['E21']
    e21.value = processing_fee
    e21.number_format = acct_fmt
    e21.border = Border(top=hair, bottom=thin, right=thin)

    e22 = ws1['E22']
    e22.value = '=SUM(E19:E21)'
    e22.font = Font(bold=True)
    e22.number_format = acct_fmt
    e22.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Estimated Closing Costs Table
    c24 = ws1['C24']
    c24.value = 'ESTIMATED CLOSING COSTS'
    c24.font = Font(bold=True)
    c24.border = Border(bottom=thin)

    c25 = ws1['C25']
    c25.value = 'Settlement Fee'
    c25.border = Border(top=thin, bottom=hair, left=thin)

    d25 = ws1['D25']
    d25.value = 'Commonly Used Fee'
    d25.border = Border(top=thin, bottom=hair)

    e25 = ws1['E25']
    e25.value = settlement_fee
    e25.number_format = acct_fmt
    e25.border = Border(top=thin, bottom=hair, right=thin)

    c26 = ws1['C26']
    c26.value = 'Deed Preparation'
    c26.border = Border(top=hair, bottom=hair, left=thin)

    d26 = ws1['D26']
    d26.value = 'Commonly Used Fee'
    d26.border = Border(top=hair, bottom=hair)

    e26 = ws1['E26']
    e26.value = deed_preparation_fee
    e26.number_format = acct_fmt
    e26.border = Border(top=hair, bottom=hair, right=thin)

    c27 = ws1['C27']
    c27.value = 'Release of Liens / Trusts'
    c27.border = Border(top=hair, bottom=thin, left=thin)

    d27 = ws1['D27']
    d27.value = 'Commonly Used Fee * Qty of Trusts or Liens'
    d27.border = Border(top=hair, bottom=thin)

    e27 = ws1['E27']
    e27.value = lien_release_fee * lien_trust_qty
    e27.number_format = acct_fmt
    e27.border = Border(top=hair, bottom=thin, right=thin)

    e28 = ws1['E28']
    e28.value = '=SUM(E25:E27)'
    e28.font = Font(bold=True)
    e28.number_format = acct_fmt
    e28.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Misc Costs Table
    c30 = ws1['C30']
    c30.value = 'MISCELLANEOUS COSTS'
    c30.font = Font(bold=True)
    c30.border = Border(bottom=thin)

    c31 = ws1['C31']
    c31.value = 'Recording Release(s)'
    c31.border = Border(top=thin, bottom=hair, left=thin)

    d31 = ws1['D31']
    d31.value = 'Common Fee * Qty of Trusts / Liens Released'
    d31.border = Border(top=thin, bottom=hair)

    e31 = ws1['E31']
    e31.value = recording_release_fee * recording_release_qty
    e31.number_format = acct_fmt
    e31.border = Border(top=thin, bottom=hair, right=thin)

    c32 = ws1['C32']
    c32.value = 'Grantor\'s Tax'
    c32.border = Border(top=hair, bottom=hair, left=thin)

    d32 = ws1['D32']
    d32.value = '% of Recommended List Price'
    d32.border = Border(top=hair, bottom=hair)

    e32 = ws1['E32']
    e32.value = grantors_tax_pct * list_price
    e32.number_format = acct_fmt
    e32.border = Border(top=hair, bottom=hair, right=thin)

    c33 = ws1['C33']
    c33.value = 'Congestion Relief Tax'
    c33.border = Border(top=hair, bottom=hair, left=thin)

    d33 = ws1['D33']
    d33.value = '% of Recommended List Price'
    d33.border = Border(top=hair, bottom=hair)

    e33 = ws1['E33']
    e33.value = congestion_tax_pct * list_price
    e33.number_format = acct_fmt
    e33.border = Border(top=hair, bottom=hair, right=thin)

    c34 = ws1['C34']
    c34.value = 'Pest Inspection'
    c34.border = Border(top=hair, bottom=hair, left=thin)

    d34 = ws1['D34']
    d34.value = 'Commonly Used Fee'
    d34.border = Border(top=hair, bottom=hair)

    e34 = ws1['E34']
    e34.value = pest_inspection_fee
    e34.number_format = acct_fmt
    e34.border = Border(top=hair, bottom=hair, right=thin)

    c35 = ws1['C35']
    c35.value = 'POA / Condo Disclosures'
    c35.border = Border(top=hair, bottom=hair, left=thin)

    d35 =ws1['D35']
    d35.value = 'Commonly Used Fee'
    d35.border = Border(top=hair, bottom=hair)

    e35 = ws1['E35']
    e35.value = poa_condo_disclosure_fee
    e35.number_format = acct_fmt
    e35.border = Border(top=hair, bottom=hair, right=thin)

    c36 = ws1['C36']
    c36.value = other_fee_name
    c36.border = Border(top=hair, bottom=thin, left=thin)

    d36 = ws1['D36']
    d36.border = Border(top=hair, bottom=thin)

    e36 = ws1['E36']
    e36.value = other_fee_amt
    e36.number_format = acct_fmt
    e36.border = Border(top=hair, bottom=thin, right=thin)

    e37 = ws1['E37']
    e37.value = '=SUM(E31:E36)'
    e37.font = Font(bold=True)
    e37.number_format = acct_fmt
    e37.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Subtotals section
    ws1.merge_cells('C39:D39')
    top_left_cell_four = ws1['C39']
    top_left_cell_four.value = 'TOTAL ESTIMATED COST OF SETTLEMENT'
    top_left_cell_four.font = Font(bold=True)
    top_left_cell_four.alignment = Alignment(horizontal='right')

    e39 = ws1['E39']
    e39.value = '=SUM(E16,E22,E28,E37)'
    e39.font = Font(bold=True)
    e39.number_format = acct_fmt
    e39.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws1.merge_cells('C41:D41')
    top_left_cell_five = ws1['C41']
    top_left_cell_five.value = 'Sales Price'
    top_left_cell_five.font = Font(bold=True)
    top_left_cell_five.alignment = Alignment(horizontal='right')

    e41 = ws1['E41']
    e41.value = list_price
    e41.font = Font(bold=True)
    e41.number_format = acct_fmt
    e41.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws1.merge_cells('C42:D42')
    top_left_cell_six = ws1['C42']
    top_left_cell_six.value = 'LESS: Total Estimated Cost of Settlement'
    top_left_cell_six.font = Font(bold=True)
    top_left_cell_six.alignment = Alignment(horizontal='right')

    e42 = ws1['E42']
    e42.value = '=-SUM(E16,E22,E28,E37)'
    e42.font = Font(bold=True)
    e42.number_format = acct_fmt
    e42.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws1.merge_cells('C44:D44')
    top_left_cell_seven = ws1['C44']
    top_left_cell_seven.value = 'ESTIMATED TOTAL NET PROCEEDS'
    top_left_cell_seven.font = Font(bold=True)
    top_left_cell_seven.alignment = Alignment(horizontal='right')

    e44 = ws1['E44']
    e44.value = '=SUM(E41,E42)'
    e44.font = Font(bold=True)
    e44.fill = PatternFill(start_color=yellow_fill, end_color=yellow_fill, fill_type='solid')
    e44.number_format = acct_fmt
    e44.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Signature Block
    c51 = ws1['C46']
    c51.value = 'PREPARED BY:'

    c52 = ws1['C47']
    c52.value = agent

    d51 = ws1['D46']
    d51.value = 'SELLER'

    d52 = ws1['D47']
    d52.value = seller

    # Freedom Logo
    freedom_logo = drawing.image.Image('freedom_logo.png')
    freedom_logo.height = 80
    freedom_logo.width = 115
    freedom_logo.anchor = 'C48'
    ws1.add_image(freedom_logo)

    # Disclosure Statement
    ws1.merge_cells('C52:E55')
    top_left_cell_eight = ws1['C52']
    top_left_cell_eight.value = '''These estimates are not guaranteed and may not include escrows. Escrow balances are reimbursed by the existing lender. Taxes, rents & association dues are pro-rated at settlement. Under Virginia Law, the seller's proceeds may not be available for up to 2 business days following recording of the deed. Seller acknowledges receipt of this statement.'''
    top_left_cell_eight.font = Font(italic=True)
    top_left_cell_eight.alignment = Alignment(wrapText=True)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        data = BytesIO(tmp.read())

    return data

    # wb.save(filename=dest_filename)